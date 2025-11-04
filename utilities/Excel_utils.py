import openpyxl


def GetRowCount(file,sheet_name):
    book=openpyxl.load_workbook(file)
    sheet=book[sheet_name]
    return sheet.max_row


def read_data(file,sheet_name,row_no,col_no):
    book = openpyxl.load_workbook(file)
    sheet = book[sheet_name]
    return sheet.cell(row=row_no,column=col_no).value


def write_data(file,sheet_name,row_no,col_no,data):
    book = openpyxl.load_workbook(file)
    sheet = book[sheet_name]
    sheet.cell(row=row_no,column=col_no).value=data
    book.save(file)


